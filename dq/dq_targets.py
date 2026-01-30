import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

def analyze_excel(file_path):
    """
    Perform basic analysis on an Excel file
    
    Args:
        file_path: Path to the Excel file
    """
    # Read Excel file from 'Targets' sheet, starting at row 8 (header at row 7)
    df = pd.read_excel(file_path, sheet_name='Targets', header=7)
    
    # Select only the columns of interest
    columns_of_interest = [
        'Document ID',
        'Country Code',
        'Version number',
        'Target area',
        'Target scope',
        'GHG target?',
        'Target type',
        'Conditionality',
        'Target Year',
        'Content',
        'Page Number'
    ]
    
    # Filter to only include these columns (handles case variations)
    available_columns = [col for col in columns_of_interest if col in df.columns]
    df = df[available_columns]
    
    print(f"Analysis of: {file_path}")
    print("=" * 50)
    
    # Basic information
    print("\n1. Dataset Overview:")
    print(f"   - Rows: {len(df)}")
    print(f"   - Columns: {len(df.columns)}")
    print(f"   - Column names: {list(df.columns)}")
    
    # Data types
    print("\n2. Data Types:")
    print(df.dtypes)
    
    # Missing values
    print("\n3. Missing Values:")
    missing = df.isnull().sum()
    missing_pct = (missing / len(df)) * 100
    missing_df = pd.DataFrame({
        'Missing Count': missing,
        'Percentage': missing_pct
    })
    print(missing_df[missing_df['Missing Count'] > 0])
    
    # Descriptive statistics
    print("\n4. Descriptive Statistics:")
    print(df.describe(include='all'))
    
    # Duplicate rows
    print("\n5. Duplicate Rows:")
    duplicates = df.duplicated().sum()
    print(f"   - Number of duplicates: {duplicates}")
    
    # Value counts for categorical columns
    print("\n6. Value Counts for Key Columns:")
    categorical_cols = ['Version number', 'Target area', 'Target scope', 'Target type', 'Conditionality', 'GHG target?']
    for col in categorical_cols:
        if col in df.columns:
            print(f"\n   {col}:")
            value_counts = df[col].value_counts()
            if len(value_counts) > 10:
                print(value_counts.head(10))
                print(f"   ... and {len(value_counts) - 10} more values")
            else:
                print(value_counts)
    
    # Check if Target Year is in Content
    print("\n7. Target Year Validation:")
    year_in_content = check_year_in_content(df)
    print(f"   - Rows where Target Year is NOT in Content: {year_in_content['missing'].sum()}")
    print(f"   - Percentage: {(year_in_content['missing'].sum() / len(df)) * 100:.2f}%")
    print(f"   - Rows with suggested alternative years: {(year_in_content['suggested_years'].notna()).sum()}")
    
    return df

def extract_years_from_text(text, future_only=False):
    """
    Extract all 4-digit years from text
    
    Args:
        text: String to extract years from
        future_only: If True, only return years >= 2015
    
    Returns:
        List of years found
    """
    if pd.isna(text):
        return []
    
    # Match complete 4-digit years (2015-2099)
    years = re.findall(r'\b(20\d{2})\b', str(text))
    
    if future_only:
        # Filter to only include years from 2015 onwards
        years = [year for year in years if int(year) >= 2015]
    
    return years

def check_year_in_content(df):
    """
    Check if Target Year appears in the Content column
    
    Args:
        df: DataFrame with 'Target Year' and 'Content' columns
    
    Returns:
        DataFrame with validation results
    """
    results = pd.DataFrame()
    
    if 'Target Year' in df.columns and 'Content' in df.columns:
        def year_exists_in_content(row):
            if pd.isna(row['Target Year']) or pd.isna(row['Content']):
                return None, None  # Can't validate if either is missing
            
            # Convert Target Year to string and extract all 4-digit years
            target_year_str = str(row['Target Year'])
            content_str = str(row['Content'])
            
            # Extract years from target year (handles ranges like "2030", "2030-2035", etc.)
            years_in_target = extract_years_from_text(target_year_str)
            
            # Check if any of these years appear in content
            for year in years_in_target:
                if year in content_str:
                    return True, None
            
            # If year not found, extract future years from content as suggestions
            years_in_content = extract_years_from_text(content_str, future_only=True)
            suggested_years = ', '.join(sorted(set(years_in_content))) if years_in_content else None
            
            return False, suggested_years
        
        validation_results = df.apply(year_exists_in_content, axis=1, result_type='expand')
        results['year_in_content'] = validation_results[0]
        results['suggested_years'] = validation_results[1]
        results['missing'] = results['year_in_content'] == False
        results['cannot_validate'] = results['year_in_content'].isna()
    
    return results

def export_analysis_report(df, output_path):
    """
    Export analysis results to Excel with multiple sheets
    
    Args:
        df: DataFrame to analyze
        output_path: Path for output Excel file
    """
    # Create a copy with missing value markers
    df_marked = df.copy()
    
    # Check year in content
    year_check = check_year_in_content(df)
    df_marked['Year_In_Content'] = year_check['year_in_content'].map({
        True: '✓', 
        False: '❌ YEAR NOT FOUND', 
        None: 'N/A'
    })
    df_marked['Suggested_Years'] = year_check['suggested_years']
    
    # Add flag columns for missing values
    for col in df.columns:
        flag_col = f'{col}_MISSING'
        df_marked[flag_col] = df[col].isnull().map({True: '❌ MISSING', False: ''})
    
    # README data with sheet names without spaces
    readme_data = {
        'Sheet Name': [
            'README',
            'Data',
            'Data_with_Flags',
            'Statistics',
            'Missing_Values',
            'Year_Not_In_Content',
            'Duplicate_Rows',
            'Incomplete_Rows',
            'Country_Code',
            'Version_number',
            'Target_area',
            'Target_scope',
            'GHG_target',
            'Target_type',
            'Conditionality'
        ],
        'Description': [
            'This documentation sheet explaining all content',
            'Clean dataset without quality flags (original data only).',
            'Complete dataset with quality flags. Missing values marked with ❌ MISSING. Year validation marked with ✓ or ❌ YEAR NOT FOUND.',
            'Descriptive statistics for all columns: count, unique values, top values, frequencies, mean, std, min, max, percentiles.',
            'Summary of missing values per column with count and percentage. Helps identify data completeness issues.',
            'Rows where Target Year does not appear in Content field. Includes suggested alternative years found in content.',
            'Complete duplicate rows grouped by Duplicate_Group ID. Group numbers start from 1. Review for potential data collection errors.',
            'All rows with at least one missing value, with ❌ flags showing which fields are incomplete.',
            'Value counts for Country Code field showing entries per country.',
            'Value counts for Version number field showing NDC version distribution.',
            'Value counts for Target area field showing distribution of target areas.',
            'Value counts for Target scope field showing distribution of target scopes.',
            'Value counts for GHG target field showing greenhouse gas target classification.',
            'Value counts for Target type field showing distribution of target types.',
            'Value counts for Conditionality field showing conditional vs unconditional targets.'
        ]
    }
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. README
        readme_df = pd.DataFrame(readme_data)
        readme_df.to_excel(writer, sheet_name='README', index=False)
        
        # 2. DATA
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # 3. DATA WITH FLAGS
        df_marked.to_excel(writer, sheet_name='Data_with_Flags', index=False)
        
        # 4. STATISTICS
        df.describe(include='all').to_excel(writer, sheet_name='Statistics')
        
        # 5. MISSING VALUES
        missing = df.isnull().sum()
        missing_pct = (missing / len(df)) * 100
        missing_df = pd.DataFrame({
            'Column': df.columns,
            'Missing Count': missing.values,
            'Percentage': missing_pct
        })
        missing_df.to_excel(writer, sheet_name='Missing_Values', index=False)
        
        # 6. YEAR VALIDATION ISSUES
        year_issues = df_marked[df_marked['Year_In_Content'] == '❌ YEAR NOT FOUND'].copy()
        if not year_issues.empty:
            cols_to_keep = [col for col in year_issues.columns if not col.endswith('_MISSING')]
            year_issues = year_issues[cols_to_keep]
            
            # Reorder to show suggested years prominently after Target Year
            if 'Suggested_Years' in cols_to_keep and 'Target Year' in cols_to_keep:
                cols_to_keep.remove('Suggested_Years')
                if 'Year_In_Content' in cols_to_keep:
                    cols_to_keep.remove('Year_In_Content')
                target_year_idx = cols_to_keep.index('Target Year')
                cols_to_keep.insert(target_year_idx + 1, 'Suggested_Years')
                cols_to_keep.insert(target_year_idx + 2, 'Year_In_Content')
            
            year_issues = year_issues[cols_to_keep]
            year_issues.to_excel(writer, sheet_name='Year_Not_In_Content', index=False)
        
        # 7. DUPLICATE ROWS
        duplicates = df[df.duplicated(keep=False)].copy()
        if not duplicates.empty:
            duplicates['Duplicate_Group'] = duplicates.groupby(list(df.columns), dropna=False).ngroup() + 1
            duplicates = duplicates.sort_values(by=['Duplicate_Group'] + list(df.columns))
            cols = ['Duplicate_Group'] + [col for col in df.columns]
            duplicates = duplicates[cols]
            duplicates.to_excel(writer, sheet_name='Duplicate_Rows', index=False)
        
        # 8. INCOMPLETE ROWS
        rows_with_missing = df[df.isnull().any(axis=1)].copy()
        if not rows_with_missing.empty:
            for col in rows_with_missing.columns:
                flag_col = f'{col}_MISSING'
                rows_with_missing[flag_col] = rows_with_missing[col].isnull().map({True: '❌', False: ''})
            rows_with_missing.to_excel(writer, sheet_name='Incomplete_Rows', index=False)
        
        # 9+. VALUE COUNTS
        categorical_cols_mapping = {
            'Country Code': 'Country_Code',
            'Version number': 'Version_number',
            'Target area': 'Target_area',
            'Target scope': 'Target_scope',
            'GHG target?': 'GHG_target',
            'Target type': 'Target_type',
            'Conditionality': 'Conditionality'
        }
        
        for col, sheet_name in categorical_cols_mapping.items():
            if col in df.columns:
                value_counts = df[col].value_counts(dropna=False).reset_index()
                value_counts.columns = [col, 'Count']
                value_counts.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get the workbook to apply formatting and hyperlinks
        workbook = writer.book
        
        # Add hyperlinks to README sheet
        readme_ws = workbook['README']
        for row_idx in range(2, len(readme_data['Sheet Name']) + 2):
            sheet_name = readme_ws[f'A{row_idx}'].value
            if sheet_name and sheet_name in workbook.sheetnames:
                cell = readme_ws[f'A{row_idx}']
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.font = Font(name='Calibri', size=11, color='0563C1', underline='single')
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Format data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                if row_idx % 2 == 0:
                    fill = alt_fill
                else:
                    fill = PatternFill()
                
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.fill = fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze top row
            ws.freeze_panes = 'A2'
            
            # Apply filters to header row
            ws.auto_filter.ref = ws.dimensions
        
        # Special formatting for README sheet
        readme_ws = workbook['README']
        readme_ws.column_dimensions['A'].width = 25
        readme_ws.column_dimensions['B'].width = 100
        
        readme_header_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        for cell in readme_ws[1]:
            cell.fill = readme_header_fill
        
        # Special formatting for error/warning sheets
        warning_sheets = ['Year_Not_In_Content', 'Duplicate_Rows', 'Incomplete_Rows']
        warning_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
        
        for sheet_name in warning_sheets:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.fill = warning_fill
        
        # Special formatting for Missing Values sheet
        if 'Missing_Values' in workbook.sheetnames:
            mv_ws = workbook['Missing_Values']
            missing_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
            for cell in mv_ws[1]:
                cell.fill = missing_fill
    
    print(f"\nReport exported to: {output_path}")

if __name__ == "__main__":
    # Specify your Excel file path
    input_file = "dq/excel/NDC-Database-Analysis_current_NEW.xlsx"
    output_file = "dq/analysis_report_targets.xlsx"
    
    # Run analysis
    data = analyze_excel(input_file)
    
    # Export report
    export_analysis_report(data, output_file)