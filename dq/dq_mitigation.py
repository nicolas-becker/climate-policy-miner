import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
import traceback

def load_definitions(file_path):
    """
    Load the hierarchical definitions from Def_Indicators sheet
    
    Args:
        file_path: Path to the Excel file
    
    Returns:
        DataFrame with Category, Purpose, Parameter mappings
    """
    try:
        # Read the definitions sheet
        df_def = pd.read_excel(file_path, sheet_name='Def_Indicators', header=8)
        
        print(f"\nLoaded definitions - Initial state:")
        print(f"   - Total rows: {len(df_def)}")
        print(f"   - Columns: {list(df_def.columns)}")
        print(f"\nFirst 10 rows of raw data:")
        print(df_def.head(10))
        
        # Remove completely empty rows
        df_def = df_def.dropna(how='all')
        print(f"\nAfter removing completely empty rows: {len(df_def)} rows")
        
        # Show first few rows with the relevant columns
        if 'Category' in df_def.columns and 'Purpose' in df_def.columns and 'Parameter' in df_def.columns:
            print(f"\nCategory-Purpose-Parameter columns (first 20 rows):")
            print(df_def[['Category', 'Purpose', 'Parameter']].head(20))
            
            # Check for non-null values
            print(f"\nNon-null counts:")
            print(f"   - Category: {df_def['Category'].notna().sum()}")
            print(f"   - Purpose: {df_def['Purpose'].notna().sum()}")
            print(f"   - Parameter: {df_def['Parameter'].notna().sum()}")
        else:
            print(f"\nWARNING: Expected columns not found!")
            print(f"Available columns: {list(df_def.columns)}")
        
        return df_def
    except Exception as e:
        print(f"\nWarning: Could not load Def_Indicators sheet: {e}")
        traceback.print_exc()
        return None

def create_clean_taxonomy(df_def):
    """
    Create a clean taxonomy table from the definitions
    
    Args:
        df_def: DataFrame with hierarchical definitions
    
    Returns:
        DataFrame with cleaned Category-Purpose-Parameter mappings
    """
    if df_def is None:
        print("No definitions loaded - cannot create taxonomy")
        return None
    
    # Check if required columns exist
    required_cols = ['Category', 'Purpose', 'Parameter']
    missing_cols = [col for col in required_cols if col not in df_def.columns]
    
    if missing_cols:
        print(f"\nERROR: Missing required columns: {missing_cols}")
        print(f"Available columns: {list(df_def.columns)}")
        return None
    
    # Filter to only rows with all three fields populated
    taxonomy = df_def[['Category', 'Purpose', 'Parameter']].copy()
    
    print(f"\nBefore processing - rows with data:")
    print(f"   - Total rows: {len(taxonomy)}")
    
    # Remove rows where all three fields are null
    taxonomy = taxonomy.dropna(how='all')
    print(f"   - After removing all-null rows: {len(taxonomy)}")
    
    # Step 1: Forward-fill Category column only
    print(f"\nStep 1: Forward-filling Category column")
    print(f"   - Category NaN count before: {taxonomy['Category'].isna().sum()}")
    taxonomy['Category'] = taxonomy['Category'].ffill()
    print(f"   - Category NaN count after: {taxonomy['Category'].isna().sum()}")
    
    # Step 2: Remove rows where both Purpose and Parameter are NaN
    print(f"\nStep 2: Removing rows where both Purpose and Parameter are NaN")
    rows_before = len(taxonomy)
    taxonomy = taxonomy[~(taxonomy['Purpose'].isna() & taxonomy['Parameter'].isna())]
    rows_removed = rows_before - len(taxonomy)
    print(f"   - Rows removed: {rows_removed}")
    print(f"   - Remaining rows: {len(taxonomy)}")
    
    # Step 3: Forward-fill Purpose and Parameter columns
    print(f"\nStep 3: Forward-filling Purpose and Parameter columns")
    print(f"   - Purpose NaN count before: {taxonomy['Purpose'].isna().sum()}")
    print(f"   - Parameter NaN count before: {taxonomy['Parameter'].isna().sum()}")
    taxonomy['Purpose'] = taxonomy['Purpose'].ffill()
    taxonomy['Parameter'] = taxonomy['Parameter'].ffill()
    print(f"   - Purpose NaN count after: {taxonomy['Purpose'].isna().sum()}")
    print(f"   - Parameter NaN count after: {taxonomy['Parameter'].isna().sum()}")
    
    # Final cleanup: Remove any remaining rows with NaN (shouldn't be any, but just in case)
    taxonomy = taxonomy.dropna(subset=['Category', 'Purpose', 'Parameter'])
    print(f"   - After final NaN removal: {len(taxonomy)} rows")
    
    if taxonomy.empty:
        print("\nWARNING: Taxonomy is empty after filtering!")
        print("Sample of original data:")
        print(df_def[['Category', 'Purpose', 'Parameter']].head(30))
        return None
    
    # Clean whitespace
    taxonomy['Category'] = taxonomy['Category'].str.strip()
    taxonomy['Purpose'] = taxonomy['Purpose'].str.strip()
    taxonomy['Parameter'] = taxonomy['Parameter'].str.strip()
    
    # Remove duplicates
    before_dedup = len(taxonomy)
    taxonomy = taxonomy.drop_duplicates()
    print(f"\nDuplicate removal:")
    print(f"   - Before: {before_dedup} rows")
    print(f"   - After: {len(taxonomy)} rows (removed {before_dedup - len(taxonomy)})")
    
    # Sort for readability
    taxonomy = taxonomy.sort_values(['Category', 'Purpose', 'Parameter'])
    
    # Reset index
    taxonomy = taxonomy.reset_index(drop=True)
    
    # Add a unique ID for each combination
    taxonomy.insert(0, 'Taxonomy_ID', range(1, len(taxonomy) + 1))
    
    print(f"\nClean Taxonomy created:")
    print(f"   - Total valid combinations: {len(taxonomy)}")
    print(f"   - Unique Categories: {taxonomy['Category'].nunique()}")
    print(f"   - Unique Purposes: {taxonomy['Purpose'].nunique()}")
    print(f"   - Unique Parameters: {taxonomy['Parameter'].nunique()}")
    
    # Print category breakdown
    print("\nCombinations per Category:")
    category_counts = taxonomy['Category'].value_counts().sort_index()
    for cat, count in category_counts.items():
        print(f"   - {cat}: {count} combinations")
    
    # Show sample of taxonomy
    print("\nSample taxonomy entries (first 10):")
    print(taxonomy.head(10))
    
    return taxonomy

def check_hierarchy_consistency(df, taxonomy):
    """
    Check if Category-Purpose-Parameter combinations are valid according to taxonomy
    
    Args:
        df: DataFrame with mitigation data
        taxonomy: Clean taxonomy DataFrame
    
    Returns:
        DataFrame with inconsistency flags
    """
    if taxonomy is None or taxonomy.empty:
        print("No taxonomy available - skipping consistency check")
        return pd.DataFrame()
    
    results = pd.DataFrame()
    
    # Build valid combinations set from taxonomy
    valid_combinations = set()
    for _, row in taxonomy.iterrows():
        combo = (
            str(row['Category']),
            str(row['Purpose']),
            str(row['Parameter'])
        )
        valid_combinations.add(combo)
    
    print(f"\nUsing {len(valid_combinations)} valid combinations for consistency check")
    
    # Check each row in data
    def check_combination(row):
        if pd.isna(row['Category']) or pd.isna(row['Purpose']) or pd.isna(row['Parameter']):
            return None  # Cannot validate if any field is missing
        
        combo = (
            str(row['Category']).strip(),
            str(row['Purpose']).strip(),
            str(row['Parameter']).strip()
        )
        
        return combo in valid_combinations
    
    if 'Category' in df.columns and 'Purpose' in df.columns and 'Parameter' in df.columns:
        results['is_valid'] = df.apply(check_combination, axis=1)
        results['is_invalid'] = results['is_valid'] == False
        results['cannot_validate'] = results['is_valid'].isna()
        
        # Count issues
        invalid_count = results['is_invalid'].sum()
        cannot_validate_count = results['cannot_validate'].sum()
        valid_count = (results['is_valid'] == True).sum()
        
        print(f"\nConsistency check results:")
        print(f"   - Valid combinations: {valid_count}")
        print(f"   - Invalid combinations: {invalid_count}")
        print(f"   - Cannot validate (missing data): {cannot_validate_count}")
        
        # Show examples of invalid combinations
        if invalid_count > 0:
            invalid_combos = df[results['is_invalid']][['Category', 'Purpose', 'Parameter']].drop_duplicates()
            print(f"\n   Examples of invalid combinations found in data:")
            for idx, row in invalid_combos.head(5).iterrows():
                print(f"      • {row['Category']} → {row['Purpose']} → {row['Parameter']}")
            if len(invalid_combos) > 5:
                print(f"      ... and {len(invalid_combos) - 5} more unique invalid combinations")
    
    return results

def analyze_excel(file_path):
    """
    Perform basic analysis on an Excel file
    
    Args:
        file_path: Path to the Excel file
    """
    # Read Excel file from 'Mitigation' sheet, starting at row 8 (header at row 7)
    df = pd.read_excel(file_path, sheet_name='Mitigation', header=7)
    
    # Select only the columns of interest for Mitigation analysis
    columns_of_interest = [
        'Document ID',
        'Country Code',
        'Version number',
        'Category',
        'Purpose',
        'Parameter',
        'Quote',
        'A-S-I',
        'Activity type',
        'Status of measure',
        'Page Number'
    ]
    
    # Filter to only include these columns (handles case variations)
    available_columns = [col for col in columns_of_interest if col in df.columns]
    df = df[available_columns]
    
    print(f"Analysis of: {file_path}")
    print("=" * 50)
    
    # Basic information
    print("\n1. Dataset Overview:")
    print(f"   - Rows: {len(df)}")
    print(f"   - Columns: {len(df.columns)}")
    print(f"   - Column names: {list(df.columns)}")
    
    # Data types
    print("\n2. Data Types:")
    print(df.dtypes)
    
    # Missing values
    print("\n3. Missing Values:")
    missing = df.isnull().sum()
    missing_pct = (missing / len(df)) * 100
    missing_df = pd.DataFrame({
        'Missing Count': missing,
        'Percentage': missing_pct
    })
    print(missing_df[missing_df['Missing Count'] > 0])
    
    # Descriptive statistics
    print("\n4. Descriptive Statistics:")
    print(df.describe(include='all'))
    
    # Duplicate rows
    print("\n5. Duplicate Rows:")
    duplicates = df.duplicated().sum()
    print(f"   - Number of duplicates: {duplicates}")
    
    # Value counts for categorical columns
    print("\n6. Value Counts for Key Columns:")
    categorical_cols = ['Country Code', 'Version number', 'Category', 'Purpose', 'Parameter', 'A-S-I', 'Activity type', 'Status of measure']
    for col in categorical_cols:
        if col in df.columns:
            print(f"\n   {col}:")
            value_counts = df[col].value_counts()
            if len(value_counts) > 10:
                print(value_counts.head(10))
                print(f"   ... and {len(value_counts) - 10} more values")
            else:
                print(value_counts)
    
    # Load definitions and create taxonomy
    print("\n7. Taxonomy Creation & Hierarchy Consistency Check:")
    df_def = load_definitions(file_path)
    taxonomy = create_clean_taxonomy(df_def)
    
    return df, taxonomy

def export_analysis_report(df, taxonomy, output_path):
    """
    Export analysis results to Excel with multiple sheets
    
    Args:
        df: DataFrame to analyze
        taxonomy: Clean taxonomy DataFrame
        output_path: Path for output Excel file
    """
    
    # Check hierarchy consistency
    consistency_check = check_hierarchy_consistency(df, taxonomy)
    
    # Create a copy with missing value markers
    df_marked = df.copy()
    
    # Add consistency check results
    if not consistency_check.empty:
        df_marked['Hierarchy_Valid'] = consistency_check['is_valid'].map({
            True: '✓',
            False: '❌ INVALID COMBINATION',
            None: 'N/A'
        })
    
    # Add flag columns for missing values
    for col in df.columns:
        flag_col = f'{col}_MISSING'
        df_marked[flag_col] = df[col].isnull().map({True: '❌ MISSING', False: ''})
    
    # README data with sheet names without spaces
    readme_data = {
        'Sheet Name': [
            'README',
            'Taxonomy',
            'Data',
            'Data_with_Flags',
            'Statistics',
            'Missing_Values',
            'Invalid_Hierarchies',
            'Invalid_Combos_Summary',
            'Duplicate_Rows',
            'Incomplete_Rows',
            'Category',
            'Purpose',
            'Parameter',
            'Category-Purpose',
            'Parameter_by_Category',
            'A-S-I',
            'Activity_type',
            'Status_of_measure',
            'ASI_by_Activity',
            'Status_by_Country',
            'Country_Code',
            'Version_number'
        ],
        'Description': [
            'This documentation sheet explaining all content',
            'Reference taxonomy with valid Category-Purpose-Parameter combinations from Def_Indicators sheet. Used for hierarchy consistency validation.',
            'Clean dataset without quality flags (original data only).',
            'Complete dataset with quality flags. Missing values marked with ❌ MISSING. Hierarchy validation marked with ✓ or ❌ INVALID COMBINATION.',
            'Descriptive statistics for all columns: count, unique values, top values, frequencies, mean, std, min, max, percentiles.',
            'Summary of missing values per column with count and percentage. Helps identify data completeness issues.',
            'Rows with Category-Purpose-Parameter combinations that do not match the reference taxonomy. Indicates potential data entry errors.',
            'Aggregated view of invalid combinations showing which ones occur most frequently. Prioritize fixing high-occurrence issues.',
            'Complete duplicate rows grouped by Duplicate_Group ID. Group numbers start from 1. Review for potential data collection errors.',
            'All rows with at least one missing value, with ❌ flags showing which fields are incomplete.',
            'Value counts for Category field showing distribution across all entries.',
            'Value counts for Purpose field showing distribution across all entries.',
            'Value counts for Parameter field showing distribution across all entries.',
            'Cross-tabulation showing how many entries exist for each Category-Purpose combination.',
            'Cross-tabulation showing how many entries exist for each Category-Parameter combination.',
            'Value counts for A-S-I (Avoid-Shift-Improve) field.',
            'Value counts for Activity type field.',
            'Value counts for Status of measure field.',
            'Cross-tabulation showing relationship between A-S-I and Activity type.',
            'Cross-tabulation showing Status of measure distribution per Country.',
            'Value counts for Country Code field showing entries per country.',
            'Value counts for Version number field showing NDC version distribution.'
        ]
    }
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write sheets in order
        # 1. README (placeholder)
        readme_df = pd.DataFrame(readme_data)
        readme_df.to_excel(writer, sheet_name='README', index=False)
        
        # 2. TAXONOMY
        if taxonomy is not None and not taxonomy.empty:
            taxonomy.to_excel(writer, sheet_name='Taxonomy', index=False)
        
        # 3. DATA
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # 4. DATA WITH FLAGS (no spaces)
        df_marked.to_excel(writer, sheet_name='Data_with_Flags', index=False)
        
        # 5. STATISTICS
        df.describe(include='all').to_excel(writer, sheet_name='Statistics')
        
        # 6. MISSING VALUES (no spaces)
        missing = df.isnull().sum()
        missing_pct = (missing / len(df)) * 100
        missing_df = pd.DataFrame({
            'Column': df.columns,
            'Missing Count': missing.values,
            'Percentage': missing_pct
        })
        missing_df.to_excel(writer, sheet_name='Missing_Values', index=False)
        
        # 7-8. HIERARCHY CONSISTENCY ISSUES (no spaces)
        if not consistency_check.empty:
            invalid_rows = df_marked[df_marked['Hierarchy_Valid'] == '❌ INVALID COMBINATION'].copy()
            if not invalid_rows.empty:
                cols_to_keep = [col for col in invalid_rows.columns if not col.endswith('_MISSING')]
                invalid_rows = invalid_rows[cols_to_keep]
                invalid_rows.to_excel(writer, sheet_name='Invalid_Hierarchies', index=False)
                
                invalid_combos = invalid_rows[['Category', 'Purpose', 'Parameter']].copy()
                invalid_combos = invalid_combos.groupby(['Category', 'Purpose', 'Parameter']).size().reset_index(name='Occurrences')
                invalid_combos = invalid_combos.sort_values('Occurrences', ascending=False)
                invalid_combos.to_excel(writer, sheet_name='Invalid_Combos_Summary', index=False)
        
        # 9. DUPLICATE ROWS (no spaces)
        duplicates = df[df.duplicated(keep=False)].copy()
        if not duplicates.empty:
            duplicates['Duplicate_Group'] = duplicates.groupby(list(df.columns), dropna=False).ngroup() + 1
            duplicates = duplicates.sort_values(by=['Duplicate_Group'] + list(df.columns))
            cols = ['Duplicate_Group'] + [col for col in df.columns]
            duplicates = duplicates[cols]
            duplicates.to_excel(writer, sheet_name='Duplicate_Rows', index=False)
        
        # 10. INCOMPLETE ROWS (no spaces)
        rows_with_missing = df[df.isnull().any(axis=1)].copy()
        if not rows_with_missing.empty:
            for col in rows_with_missing.columns:
                flag_col = f'{col}_MISSING'
                rows_with_missing[flag_col] = rows_with_missing[col].isnull().map({True: '❌', False: ''})
            rows_with_missing.to_excel(writer, sheet_name='Incomplete_Rows', index=False)
        
        # 11+. VALUE COUNTS AND CROSS-TABS (replacing spaces with underscores)
        categorical_cols_mapping = {
            'Country Code': 'Country_Code',
            'Version number': 'Version_number',
            'Category': 'Category',
            'Purpose': 'Purpose',
            'Parameter': 'Parameter',
            'A-S-I': 'A-S-I',
            'Activity type': 'Activity_type',
            'Status of measure': 'Status_of_measure'
        }
        
        for col, sheet_name in categorical_cols_mapping.items():
            if col in df.columns:
                value_counts = df[col].value_counts(dropna=False).reset_index()
                value_counts.columns = [col, 'Count']
                value_counts.to_excel(writer, sheet_name=sheet_name, index=False)
        
        if 'Category' in df.columns and 'Purpose' in df.columns:
            cat_purpose = df.groupby(['Category', 'Purpose']).size().reset_index(name='Count')
            cat_purpose = cat_purpose.sort_values(['Category', 'Count'], ascending=[True, False])
            cat_purpose.to_excel(writer, sheet_name='Category-Purpose', index=False)
        
        if 'Category' in df.columns and 'Parameter' in df.columns:
            parameter_cat = df.groupby(['Category', 'Parameter']).size().reset_index(name='Count')
            parameter_cat = parameter_cat.sort_values(['Category', 'Count'], ascending=[True, False])
            parameter_cat.to_excel(writer, sheet_name='Parameter_by_Category', index=False)
        
        if 'A-S-I' in df.columns and 'Activity type' in df.columns:
            asi_activity = df.groupby(['A-S-I', 'Activity type']).size().reset_index(name='Count')
            asi_activity = asi_activity.sort_values(['A-S-I', 'Count'], ascending=[True, False])
            asi_activity.to_excel(writer, sheet_name='ASI_by_Activity', index=False)
        
        if 'Country Code' in df.columns and 'Status of measure' in df.columns:
            status_country = df.groupby(['Country Code', 'Status of measure']).size().reset_index(name='Count')
            status_country = status_country.sort_values(['Country Code', 'Count'], ascending=[True, False])
            status_country.to_excel(writer, sheet_name='Status_by_Country', index=False)
        
        # Get the workbook to apply formatting and hyperlinks
        workbook = writer.book
        
        # Add hyperlinks to README sheet
        readme_ws = workbook['README']
        for row_idx in range(2, len(readme_data['Sheet Name']) + 2):
            sheet_name = readme_ws[f'A{row_idx}'].value
            if sheet_name and sheet_name in workbook.sheetnames:
                cell = readme_ws[f'A{row_idx}']
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.font = Font(name='Calibri', size=11, color='0563C1', underline='single')
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Format data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                if row_idx % 2 == 0:
                    fill = alt_fill
                else:
                    fill = PatternFill()
                
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.fill = fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze top row
            ws.freeze_panes = 'A2'
            
            # Apply filters to header row
            ws.auto_filter.ref = ws.dimensions
        
        # Special formatting for README sheet
        readme_ws = workbook['README']
        readme_ws.column_dimensions['A'].width = 25
        readme_ws.column_dimensions['B'].width = 100
        
        readme_header_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        for cell in readme_ws[1]:
            cell.fill = readme_header_fill
        
        # Special formatting for Taxonomy sheet
        if 'Taxonomy' in workbook.sheetnames:
            tax_ws = workbook['Taxonomy']
            taxonomy_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            for cell in tax_ws[1]:
                cell.fill = taxonomy_fill
        
        # Special formatting for error/warning sheets
        warning_sheets = ['Invalid_Hierarchies', 'Invalid_Combos_Summary', 'Duplicate_Rows', 'Incomplete_Rows']
        warning_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
        
        for sheet_name in warning_sheets:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.fill = warning_fill
        
        # Special formatting for Missing Values sheet
        if 'Missing_Values' in workbook.sheetnames:
            mv_ws = workbook['Missing_Values']
            missing_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
            for cell in mv_ws[1]:
                cell.fill = missing_fill
    
    print(f"\nReport exported to: {output_path}")

if __name__ == "__main__":
    # Specify your Excel file path
    input_file = "dq/excel/NDC-Database-Analysis_current_NEW.xlsx"
    output_file = "dq/analysis_report_mitigation.xlsx"
    
    # Run analysis
    data, taxonomy = analyze_excel(input_file)
    
    # Export report
    export_analysis_report(data, taxonomy, output_file)