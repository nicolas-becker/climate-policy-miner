import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import os
import sys
import time
from tqdm import tqdm
from dotenv import load_dotenv

# Add src directory to path for imports
src_path = Path(__file__).resolve().parent.parent / 'src'
sys.path.insert(0, str(src_path))

from classification_utils import (
    QuoteTypeObject,
    MeasureTypeObject,
    AdaptationObject_FewShot,
    zero_shot_tagger,
    measure_mapping
)
from langchain_openai import AzureChatOpenAI
from langchain_community.callbacks import get_openai_callback

# Load environment variables
env_path = Path(__file__).resolve().parent.parent / '.env'
load_dotenv(dotenv_path=env_path)

# Initialize LLM
LLM = AzureChatOpenAI(
    openai_api_version=os.environ["AZURE_OPENAI_API_VERSION"],
    azure_deployment=os.environ["AZURE_OPENAI_CHAT_DEPLOYMENT_NAME"],
    model_version='2024-07-18',
    temperature=0
)

def validate_measure_classification(df, llm, max_retries=3):
    """
    Validate adaptation measures by AI-classifying the Quote and comparing with Parameter.
    Only processes NDC 3.X entries and each unique quote once.
    
    Args:
        df: DataFrame with measures data
        llm: Language model for classification
        max_retries: Maximum retries for API errors
    
    Returns:
        DataFrame with validation results
    """
    print("\n" + "="*80)
    print("ADAPTATION MEASURES VALIDATION - AI CLASSIFICATION")
    print("="*80)
    
    # Filter to NDC 3.X only
    ndc3_mask = df['Version number'].str.contains('NDC 3', case=False, na=False)
    df_ndc3 = df[ndc3_mask].copy()
    
    print(f"\nFiltered to NDC 3.X entries:")
    print(f"   - Total rows in dataset: {len(df)}")
    print(f"   - NDC 3.X rows: {len(df_ndc3)}")
    print(f"   - Rows excluded: {len(df) - len(df_ndc3)}")
    
    if len(df_ndc3) == 0:
        print("\n⚠️  No NDC 3.X entries found!")
        return pd.DataFrame()
    
    # Get unique quotes to avoid re-classification
    unique_quotes = df_ndc3['Quote'].dropna().unique()
    print(f"\nUnique quotes to classify: {len(unique_quotes)}")
    
    # Initialize results storage
    classification_cache = {}
    total_tokens = 0
    
    print("\nStarting AI classification...")
    print(f"This may take approximately {len(unique_quotes) * 2 / 60:.1f} minutes")
    print("(3 classification steps per quote)")
    
    # Classify each unique quote once
    for quote in tqdm(unique_quotes, desc="Classifying measures"):
        retry_count = 0
        success = False
        
        while retry_count < max_retries and not success:
            try:
                # Step 1: Check if it's a measure
                with get_openai_callback() as cb:
                    quote_type = zero_shot_tagger(quote, llm, QuoteTypeObject)
                    total_tokens += cb.total_tokens
                
                if quote_type.measure != 'True':
                    # Not a measure at all
                    classification_cache[quote] = {
                        'is_measure': False,
                        'is_adaptation_measure': False,
                        'predicted_parameters': [],
                        'raw_attributes': {}
                    }
                    success = True
                    continue
                
                # Step 2: Check if it's an adaptation measure
                with get_openai_callback() as cb:
                    measure_type = zero_shot_tagger(quote, llm, MeasureTypeObject)
                    total_tokens += cb.total_tokens
                
                if measure_type.adaptation_measure != 'True':
                    # It's a measure but not adaptation
                    classification_cache[quote] = {
                        'is_measure': True,
                        'is_adaptation_measure': False,
                        'predicted_parameters': [],
                        'raw_attributes': {}
                    }
                    success = True
                    continue
                
                # Step 3: Classify adaptation measure type
                all_attributes = {}
                
                with get_openai_callback() as cb:
                    adapt_attrs = zero_shot_tagger(quote, llm, AdaptationObject_FewShot)
                    total_tokens += cb.total_tokens
                
                all_attributes.update({
                    'R_Infrares': adapt_attrs.R_Infrares,
                    'R_System': adapt_attrs.R_System,
                    'R_Risk': adapt_attrs.R_Risk,
                    'R_Tech': adapt_attrs.R_Tech,
                    'R_Monitoring': adapt_attrs.R_Monitoring,
                    'R_Emergency': adapt_attrs.R_Emergency,
                    'R_Education': adapt_attrs.R_Education,
                    'R_Planning': adapt_attrs.R_Planning,
                    'R_Relocation': adapt_attrs.R_Relocation,
                    'R_Redundancy': adapt_attrs.R_Redundancy,
                    'R_Laws': adapt_attrs.R_Laws,
                    'R_Design': adapt_attrs.R_Design,
                    'R_Other': adapt_attrs.R_Other
                })
                
                # Add missing attributes from old model (commented out in FewShot)
                missing_attributes = {
                    'R_Maintain': 'False',
                    'R_Inform': 'False',
                    'R_Warning': 'False',
                    'R_Disinvest': 'False'
                }
                
                # Add all mitigation attributes with False (since measure_mapping expects them)
                mitigation_attributes = {
                    'A_Complan': 'False', 'A_Natmobplan': 'False', 'A_SUMP': 'False',
                    'A_LATM': 'False', 'A_Landuse': 'False', 'A_Density': 'False',
                    'A_Mixuse': 'False', 'A_Urban': 'False', 'S_Infraimprove': 'False',
                    'S_Infraexpansion': 'False', 'S_Intermodality': 'False',
                    'I_Freighteff': 'False', 'I_Load': 'False', 'S_Railfreight': 'False',
                    'I_Education': 'False', 'I_Ecodriving': 'False', 'I_Capacity': 'False',
                    'I_Campaigns': 'False', 'A_TDM': 'False', 'S_Parking': 'False',
                    'A_Parkingprice': 'False', 'A_Caraccess': 'False', 'A_Commute': 'False',
                    'A_Work': 'False', 'A_Teleworking': 'False', 'A_Economic': 'False',
                    'A_Emistrad': 'False', 'A_Finance': 'False', 'A_Procurement': 'False',
                    'A_Fossilfuelsubs': 'False', 'A_Fueltax': 'False', 'A_Vehicletax': 'False',
                    'A_Roadcharging': 'False', 'S_PublicTransport': 'False',
                    'S_PTIntegration': 'False', 'S_PTPriority': 'False', 'S_BRT': 'False',
                    'S_Activemobility': 'False', 'S_Walking': 'False', 'S_Cycling': 'False',
                    'S_Sharedmob': 'False', 'S_Ondemand': 'False', 'S_Maas': 'False',
                    'I_Other': 'False', 'I_ITS': 'False', 'I_Autonomous': 'False',
                    'I_DataModelling': 'False', 'I_Vehicleimprove': 'False',
                    'I_Fuelqualimprove': 'False', 'I_Inspection': 'False',
                    'I_Efficiencystd': 'False', 'I_Vehicleeff': 'False', 'A_LEZ': 'False',
                    'I_VehicleRestrictions': 'False', 'I_Vehiclescrappage': 'False',
                    'I_Lowemissionincentive': 'False', 'I_Altfuels': 'False',
                    'I_Ethanol': 'False', 'I_Biofuel': 'False', 'I_LPGCNGLNG': 'False',
                    'I_Hydrogen': 'False', 'I_RE': 'False', 'I_Transportlabel': 'False',
                    'I_Efficiencylabel': 'False', 'I_Freightlabel': 'False',
                    'I_Vehiclelabel': 'False', 'I_Fuellabel': 'False', 'I_Emobility': 'False',
                    'I_Emobilitycharging': 'False', 'I_Smartcharging': 'False',
                    'I_Emobilitypurchase': 'False', 'I_ICEdiesel': 'False',
                    'S_Micromobility': 'False', 'I_Aviation': 'False', 'I_Aircraftfleet': 'False',
                    'I_CO2certificate': 'False', 'I_Capacityairport': 'False',
                    'I_Jetfuel': 'False', 'I_Airtraffic': 'False', 'I_Shipping': 'False',
                    'I_Onshorepower': 'False', 'I_PortInfra': 'False', 'I_Shipefficiency': 'False',
                    'A_LUP': 'False', 'A_TOD': 'False'
                }
                
                # Add missing attributes
                for key, value in {**missing_attributes, **mitigation_attributes}.items():
                    if key not in all_attributes:
                        all_attributes[key] = value
                
                # Step 4: Apply mapping function to get list of parameters
                temp_series = pd.Series(all_attributes)
                predicted_params = measure_mapping(temp_series)
                
                classification_cache[quote] = {
                    'is_measure': True,
                    'is_adaptation_measure': True,
                    'predicted_parameters': predicted_params,
                    'raw_attributes': all_attributes
                }
                
                success = True
                
            except Exception as e:
                retry_count += 1
                if "rate" in str(e).lower():
                    wait_time = (2 ** retry_count) + (retry_count * 0.5)
                    print(f"\n⚠️  Rate limit hit. Waiting {wait_time:.1f}s before retry {retry_count}/{max_retries}...")
                    time.sleep(wait_time)
                else:
                    print(f"\n❌ Error classifying quote: {e}")
                    print(f"   Quote preview: {quote[:100]}...")
                    if retry_count >= max_retries:
                        classification_cache[quote] = {
                            'is_measure': None,
                            'is_adaptation_measure': None,
                            'predicted_parameters': [],
                            'raw_attributes': {},
                            'error': str(e)
                        }
                        success = True  # Mark as success to move on
    
    print(f"\n✓ Classification complete")
    print(f"   - Total API tokens used: {total_tokens:,}")
    print(f"   - Estimated cost: ${(total_tokens / 1000) * 0.045:.2f}")
    
    # Map results back to all rows (including duplicate quotes with different parameters)
    print("\nMapping results to all rows...")
    validation_results = []
    
    for idx, row in df_ndc3.iterrows():
        quote = row['Quote']
        
        if pd.isna(quote):
            validation_results.append({
                'AI_Is_Measure': None,
                'AI_Is_Adaptation': None,
                'AI_Predicted_Parameters': None,
                'Parameter_In_AI_List': None,
                'Overall_Valid': None,
                'Is_False_Positive': False
            })
            continue
        
        result = classification_cache.get(quote, {})
        
        has_parameter = pd.notna(row.get('Parameter'))
        ai_is_measure = result.get('is_measure', None)
        ai_is_adaptation = result.get('is_adaptation_measure', None)
        predicted_params = result.get('predicted_parameters', [])
        
        # Check if it's a false positive (has Parameter but AI says not adaptation measure)
        is_false_positive = has_parameter and (ai_is_adaptation == False)
        
        # Check if database parameter is in AI prediction list
        parameter_match = None
        if has_parameter and ai_is_adaptation and predicted_params:
            db_param = str(row['Parameter']).strip()
            parameter_match = db_param in predicted_params
        elif has_parameter and ai_is_adaptation and not predicted_params:
            # AI classified as adaptation but found no parameters
            parameter_match = False
        
        # Overall validation
        overall_valid = None
        if ai_is_adaptation and has_parameter:
            overall_valid = parameter_match
        
        # Format predicted parameters as string for display
        params_display = ', '.join(predicted_params) if predicted_params else '--'
        
        validation_results.append({
            'AI_Is_Measure': ai_is_measure,
            'AI_Is_Adaptation': ai_is_adaptation,
            'AI_Predicted_Parameters': params_display,
            'Parameter_In_AI_List': parameter_match,
            'Overall_Valid': overall_valid,
            'Is_False_Positive': is_false_positive
        })
    
    validation_df = pd.DataFrame(validation_results, index=df_ndc3.index)
    
    # Print summary statistics
    print("\n" + "="*80)
    print("VALIDATION SUMMARY")
    print("="*80)
    
    print(f"\nClassification Results:")
    print(f"   - AI classified as measure: {(validation_df['AI_Is_Measure'] == True).sum()}")
    print(f"   - AI classified as NOT measure: {(validation_df['AI_Is_Measure'] == False).sum()}")
    print(f"   - AI classified as adaptation measure: {(validation_df['AI_Is_Adaptation'] == True).sum()}")
    print(f"   - AI classified as NOT adaptation: {(validation_df['AI_Is_Adaptation'] == False).sum()}")
    print(f"   - False positives (has Parameter but not adaptation): {validation_df['Is_False_Positive'].sum()}")
    
    adaptation_count = (validation_df['AI_Is_Adaptation'] == True).sum()
    if adaptation_count > 0:
        print(f"\nParameter Validation (for confirmed adaptation measures):")
        print(f"   - Database parameter IN AI list (match): {(validation_df['Parameter_In_AI_List'] == True).sum()}")
        print(f"   - Database parameter NOT in AI list (mismatch): {(validation_df['Parameter_In_AI_List'] == False).sum()}")
        print(f"   - Overall valid: {(validation_df['Overall_Valid'] == True).sum()}")
        print(f"   - Overall invalid: {(validation_df['Overall_Valid'] == False).sum()}")
    
    return validation_df

def export_measures_validation_report(df_ndc3, validation_df, output_path):
    """
    Export adaptation measures validation results to Excel with formatted sheets (NDC 3.X only)
    
    Args:
        df_ndc3: DataFrame with NDC 3.X measures data
        validation_df: DataFrame with AI validation results
        output_path: Path for output Excel file
    """
    # Merge validation results with data
    df_with_validation = df_ndc3.copy()
    
    for idx in validation_df.index:
        df_with_validation.loc[idx, 'AI_Is_Measure'] = validation_df.loc[idx, 'AI_Is_Measure']
        df_with_validation.loc[idx, 'AI_Is_Adaptation'] = validation_df.loc[idx, 'AI_Is_Adaptation']
        df_with_validation.loc[idx, 'AI_Predicted_Parameters'] = validation_df.loc[idx, 'AI_Predicted_Parameters']
        df_with_validation.loc[idx, 'Parameter_In_AI_List'] = validation_df.loc[idx, 'Parameter_In_AI_List']
        df_with_validation.loc[idx, 'Overall_Valid'] = validation_df.loc[idx, 'Overall_Valid']
    
    # Apply visual indicators
    df_with_validation['AI_Is_Measure'] = df_with_validation['AI_Is_Measure'].map({
        True: '✓ Measure',
        False: '❌ Not a measure',
        None: 'N/A'
    })
    df_with_validation['AI_Is_Adaptation'] = df_with_validation['AI_Is_Adaptation'].map({
        True: '✓ Adaptation',
        False: '❌ Not adaptation',
        None: 'N/A'
    })
    df_with_validation['Parameter_In_AI_List'] = df_with_validation['Parameter_In_AI_List'].map({
        True: '✓',
        False: '❌ NOT IN LIST',
        None: 'N/A'
    })
    df_with_validation['Overall_Valid'] = df_with_validation['Overall_Valid'].map({
        True: '✓ Valid',
        False: '❌ Invalid',
        None: 'N/A'
    })
    
    # README data
    readme_data = {
        'Sheet Name': [
            'README',
            'Data',
            'Data_with_Validation',
            'Validation_Summary',
            'False_Positives',
            'Parameter_Mismatches'
        ],
        'Description': [
            'This documentation sheet explaining all content - NDC 3.X Adaptation Measures AI Validation',
            'NDC 3.X adaptation measures dataset (clean, without validation flags).',
            'NDC 3.X adaptation measures with AI validation results.',
            'Summary of AI validation showing match/mismatch statistics.',
            'Entries with Parameter but AI classified as NOT an adaptation measure (potential false positives).',
            'Rows where database Parameter is NOT in AI-predicted parameter list.'
        ]
    }
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. README
        readme_df = pd.DataFrame(readme_data)
        readme_df.to_excel(writer, sheet_name='README', index=False)
        
        # 2. DATA (clean)
        df_ndc3.to_excel(writer, sheet_name='Data', index=False)
        
        # 3. DATA WITH VALIDATION
        df_with_validation.to_excel(writer, sheet_name='Data_with_Validation', index=False)
        
        # 4. VALIDATION SUMMARY
        validation_summary = pd.DataFrame({
            'Metric': [
                'Total NDC 3.X measures validated',
                'Unique quotes classified',
                'AI classified as measure',
                'AI classified as NOT measure',
                'AI classified as adaptation measure',
                'AI classified as NOT adaptation',
                'False positives (has Parameter but not adaptation)',
                'Database parameter IN AI list (match)',
                'Database parameter NOT in AI list (mismatch)',
                'Overall valid entries',
                'Overall invalid entries'
            ],
            'Count': [
                len(validation_df),
                df_ndc3['Quote'].nunique(),
                (validation_df['AI_Is_Measure'] == True).sum(),
                (validation_df['AI_Is_Measure'] == False).sum(),
                (validation_df['AI_Is_Adaptation'] == True).sum(),
                (validation_df['AI_Is_Adaptation'] == False).sum(),
                validation_df['Is_False_Positive'].sum(),
                (validation_df['Parameter_In_AI_List'] == True).sum(),
                (validation_df['Parameter_In_AI_List'] == False).sum(),
                (validation_df['Overall_Valid'] == True).sum(),
                (validation_df['Overall_Valid'] == False).sum()
            ]
        })
        validation_summary.to_excel(writer, sheet_name='Validation_Summary', index=False)
        
        # 5. FALSE POSITIVES
        false_positive_indices = validation_df[validation_df['Is_False_Positive'] == True].index
        if len(false_positive_indices) > 0:
            false_positives = df_with_validation.loc[false_positive_indices]
            false_positives.to_excel(writer, sheet_name='False_Positives', index=False)
        
        # 6. PARAMETER MISMATCHES
        mismatch_indices = validation_df[validation_df['Parameter_In_AI_List'] == False].index
        if len(mismatch_indices) > 0:
            mismatches = df_with_validation.loc[mismatch_indices]
            priority_cols = ['Parameter', 'AI_Predicted_Parameters', 'Parameter_In_AI_List', 'Quote']
            other_cols = [c for c in mismatches.columns if c not in priority_cols]
            final_cols = [c for c in priority_cols if c in mismatches.columns] + other_cols
            mismatches[final_cols].to_excel(writer, sheet_name='Parameter_Mismatches', index=False)
        
        # Apply formatting (same as mitigation script)
        workbook = writer.book
        
        readme_ws = workbook['README']
        for row_idx in range(2, len(readme_data['Sheet Name']) + 2):
            sheet_name = readme_ws[f'A{row_idx}'].value
            if sheet_name and sheet_name in workbook.sheetnames:
                cell = readme_ws[f'A{row_idx}']
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.font = Font(name='Calibri', size=11, color='0563C1', underline='single')
        
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                fill = alt_fill if row_idx % 2 == 0 else PatternFill()
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.fill = fill
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
            
            ws.freeze_panes = 'A2'
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions
        
        readme_ws.column_dimensions['A'].width = 30
        readme_ws.column_dimensions['B'].width = 100
        readme_header_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        for cell in readme_ws[1]:
            cell.fill = readme_header_fill
        
        if 'Validation_Summary' in workbook.sheetnames:
            val_ws = workbook['Validation_Summary']
            validation_fill = PatternFill(start_color='9B59B6', end_color='9B59B6', fill_type='solid')
            for cell in val_ws[1]:
                cell.fill = validation_fill
        
        warning_sheets = ['False_Positives', 'Parameter_Mismatches']
        warning_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
        for sheet_name in warning_sheets:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.fill = warning_fill
    
    print(f"\n✓ Adaptation measures validation report exported to: {output_path}")

if __name__ == "__main__":
    input_file = "dq/excel/NDC-Database-Analysis_current_NEW.xlsx"
    output_file = "dq/ai_validation_ndc-3-X_adaptation_measures.xlsx"
    
    print("\n" + "="*80)
    print("AI PARAMETER VALIDATION - NDC 3.X ADAPTATION MEASURES")
    print("="*80)
    
    # Load data
    print(f"\nLoading data from: {input_file}")
    df = pd.read_excel(input_file, sheet_name='Adaptation', header=7)
    
    columns_of_interest = [
        'Document ID', 'Country Code', 'Version number', 
        'Parameter', 'Quote', 'Page Number'
    ]
    available_columns = [col for col in columns_of_interest if col in df.columns]
    df = df[available_columns]

    print(f"\nLoaded {len(df)} measure entries")
    
    # Filter to NDC 3.X
    ndc3_mask = df['Version number'].str.contains('NDC 3', case=False, na=False)
    df_ndc3 = df[ndc3_mask].copy()
    
    print(f"Filtered to {len(df_ndc3)} NDC 3.X measure entries")
    print(f"Unique quotes to classify: {df_ndc3['Quote'].nunique()}")
    
    # Run AI validation
    validation_results = validate_measure_classification(df, LLM)
    
    if not validation_results.empty:
        # Export results
        export_measures_validation_report(df_ndc3, validation_results, output_file)
        
        print("\n" + "="*80)
        print("AI VALIDATION COMPLETE")
        print("="*80)
    else:
        print("\n⚠️  No validation results to export")