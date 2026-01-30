import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import os
import sys
import time
from tqdm import tqdm
from dotenv import load_dotenv

# Add src directory to path for imports
src_path = Path(__file__).resolve().parent.parent / 'src'
sys.path.insert(0, str(src_path))

from classification_utils import (
    QuoteTypeObject, 
    TargetObject, 
    zero_shot_tagger,
    target_area_mapping,
    ghg_target_mapping,
    conditionality_mapping
)
from langchain_openai import AzureChatOpenAI
from langchain_community.callbacks import get_openai_callback

# Load environment variables
env_path = Path(__file__).resolve().parent.parent / '.env'
load_dotenv(dotenv_path=env_path)

# Initialize LLM
LLM = AzureChatOpenAI(
    openai_api_version=os.environ["AZURE_OPENAI_API_VERSION"],
    azure_deployment=os.environ["AZURE_OPENAI_CHAT_DEPLOYMENT_NAME"],
    model_version='2024-07-18',
    temperature=0
)

def validate_parameter_classification(df, llm, max_retries=3):
    """
    Validate the Parameter column by AI-classifying the Content and comparing results.
    Only processes NDC 3.X entries and each unique quote once.
    
    Args:
        df: DataFrame with targets data
        llm: Language model for classification
        max_retries: Maximum retries for API errors
    
    Returns:
        DataFrame with validation results
    """
    print("\n" + "="*80)
    print("PARAMETER VALIDATION - AI CLASSIFICATION")
    print("="*80)
    
    # Filter to NDC 3.X only
    ndc3_mask = df['Version number'].str.contains('NDC 3', case=False, na=False)
    df_ndc3 = df[ndc3_mask].copy()
    
    print(f"\nFiltered to NDC 3.X entries:")
    print(f"   - Total rows in dataset: {len(df)}")
    print(f"   - NDC 3.X rows: {len(df_ndc3)}")
    print(f"   - Rows excluded: {len(df) - len(df_ndc3)}")
    
    if len(df_ndc3) == 0:
        print("\n⚠️  No NDC 3.X entries found!")
        return pd.DataFrame()
    
    # Get unique quotes to avoid re-classification
    unique_quotes = df_ndc3['Content'].dropna().unique()
    print(f"\nUnique quotes to classify: {len(unique_quotes)}")
    
    # Initialize results storage
    classification_cache = {}
    total_tokens = 0
    
    print("\nStarting AI classification...")
    print(f"This may take approximately {len(unique_quotes) * 2 / 60:.1f} minutes")
    
    # Classify each unique quote once
    for quote in tqdm(unique_quotes, desc="Classifying quotes"):
        retry_count = 0
        success = False
        
        while retry_count < max_retries and not success:
            try:
                # Step 1: Check if it's a target
                with get_openai_callback() as cb:
                    quote_type = zero_shot_tagger(quote, llm, QuoteTypeObject)
                    total_tokens += cb.total_tokens
                
                if quote_type.target == 'True':
                    # Step 2: Get target attributes
                    with get_openai_callback() as cb:
                        target_attrs = zero_shot_tagger(quote, llm, TargetObject)
                        total_tokens += cb.total_tokens
                    
                    # Step 3: Apply mapping functions
                    temp_row = {
                        'energy': target_attrs.energy,
                        'transport': target_attrs.transport,
                        'sector_level': target_attrs.sector_level,
                        'mitigation': target_attrs.mitigation,
                        'adaptation': target_attrs.adaptation,
                        'ghg': target_attrs.ghg,
                        'net_zero': target_attrs.net_zero,
                        'conditional': target_attrs.conditional,
                        'unconditional': target_attrs.unconditional
                    }
                    
                    temp_series = pd.Series(temp_row)
                    
                    ai_target_area = target_area_mapping(temp_series)
                    ai_ghg = ghg_target_mapping(temp_series)
                    ai_conditionality = conditionality_mapping(temp_series)
                    
                    classification_cache[quote] = {
                        'is_target': True,
                        'target_area': ai_target_area,
                        'ghg_target': ai_ghg,
                        'conditionality': ai_conditionality,
                        'raw_attributes': temp_row
                    }
                else:
                    classification_cache[quote] = {
                        'is_target': False,
                        'target_area': '--',
                        'ghg_target': '--',
                        'conditionality': '--',
                        'raw_attributes': {}
                    }
                
                success = True
                
            except Exception as e:
                retry_count += 1
                if "rate" in str(e).lower():
                    wait_time = (2 ** retry_count) + (retry_count * 0.5)
                    print(f"\n⚠️  Rate limit hit. Waiting {wait_time:.1f}s before retry {retry_count}/{max_retries}...")
                    time.sleep(wait_time)
                else:
                    print(f"\n❌ Error classifying quote: {e}")
                    if retry_count >= max_retries:
                        classification_cache[quote] = {
                            'is_target': None,
                            'target_area': 'ERROR',
                            'ghg_target': 'ERROR',
                            'conditionality': 'ERROR',
                            'raw_attributes': {},
                            'error': str(e)
                        }
    
    print(f"\n✓ Classification complete")
    print(f"   - Total API tokens used: {total_tokens:,}")
    print(f"   - Estimated cost: ${(total_tokens / 1000) * 0.045:.2f}")
    
    # Map results back to original dataframe
    print("\nMapping results to all rows...")
    validation_results = []
    
    for idx, row in df_ndc3.iterrows():
        quote = row['Content']
        
        if pd.isna(quote):
            validation_results.append({
                'AI_Is_Target': None,
                'AI_Target_Area': None,
                'AI_GHG_Target': None,
                'AI_Conditionality': None,
                'Target_Area_Match': None,
                'GHG_Match': None,
                'Conditionality_Match': None,
                'Overall_Valid': None,
                'Is_False_Positive': False
            })
            continue
        
        result = classification_cache.get(quote, {})
        
        has_parameter = pd.notna(row.get('Target area')) or pd.notna(row.get('GHG target?')) or pd.notna(row.get('Conditionality'))
        ai_is_target = result.get('is_target', None)
        
        is_false_positive = has_parameter and (ai_is_target == False)
        
        target_area_match = None
        ghg_match = None
        conditionality_match = None
        
        if ai_is_target and has_parameter:
            if pd.notna(row.get('Target area')):
                target_area_match = (str(row['Target area']).strip() == str(result['target_area']).strip())
            
            if pd.notna(row.get('GHG target?')):
                ghg_match = (str(row['GHG target?']).strip() == str(result['ghg_target']).strip())
            
            if pd.notna(row.get('Conditionality')):
                conditionality_match = (str(row['Conditionality']).strip() == str(result['conditionality']).strip())
        
        overall_valid = None
        if ai_is_target and has_parameter:
            matches = [m for m in [target_area_match, ghg_match, conditionality_match] if m is not None]
            overall_valid = all(matches) if matches else None
        
        validation_results.append({
            'AI_Is_Target': ai_is_target,
            'AI_Target_Area': result.get('target_area'),
            'AI_GHG_Target': result.get('ghg_target'),
            'AI_Conditionality': result.get('conditionality'),
            'Target_Area_Match': target_area_match,
            'GHG_Match': ghg_match,
            'Conditionality_Match': conditionality_match,
            'Overall_Valid': overall_valid,
            'Is_False_Positive': is_false_positive
        })
    
    validation_df = pd.DataFrame(validation_results, index=df_ndc3.index)
    
    # Print summary statistics
    print("\n" + "="*80)
    print("VALIDATION SUMMARY")
    print("="*80)
    
    total_classified = (validation_df['AI_Is_Target'] == True).sum()
    false_positives = validation_df['Is_False_Positive'].sum()
    
    print(f"\nClassification Results:")
    print(f"   - AI classified as target: {total_classified}")
    print(f"   - AI classified as NOT target: {(validation_df['AI_Is_Target'] == False).sum()}")
    print(f"   - False positives (has Parameter but not a target): {false_positives}")
    
    if total_classified > 0:
        print(f"\nAttribute Validation (for confirmed targets):")
        print(f"   - Target Area matches: {(validation_df['Target_Area_Match'] == True).sum()}")
        print(f"   - Target Area mismatches: {(validation_df['Target_Area_Match'] == False).sum()}")
        print(f"   - GHG target matches: {(validation_df['GHG_Match'] == True).sum()}")
        print(f"   - GHG target mismatches: {(validation_df['GHG_Match'] == False).sum()}")
        print(f"   - Conditionality matches: {(validation_df['Conditionality_Match'] == True).sum()}")
        print(f"   - Conditionality mismatches: {(validation_df['Conditionality_Match'] == False).sum()}")
        print(f"   - Overall valid: {(validation_df['Overall_Valid'] == True).sum()}")
        print(f"   - Overall invalid: {(validation_df['Overall_Valid'] == False).sum()}")
    
    return validation_df

def export_ai_validation_report(df_ndc3, validation_df, output_path):
    """
    Export AI validation results to Excel with formatted sheets (NDC 3.X only)
    
    Args:
        df_ndc3: DataFrame with NDC 3.X targets data
        validation_df: DataFrame with AI validation results
        output_path: Path for output Excel file
    """
    # Merge validation results with data
    df_with_validation = df_ndc3.copy()
    
    for idx in validation_df.index:
        df_with_validation.loc[idx, 'AI_Is_Target'] = validation_df.loc[idx, 'AI_Is_Target']
        df_with_validation.loc[idx, 'AI_Target_Area'] = validation_df.loc[idx, 'AI_Target_Area']
        df_with_validation.loc[idx, 'AI_GHG_Target'] = validation_df.loc[idx, 'AI_GHG_Target']
        df_with_validation.loc[idx, 'AI_Conditionality'] = validation_df.loc[idx, 'AI_Conditionality']
        df_with_validation.loc[idx, 'Target_Area_Match'] = validation_df.loc[idx, 'Target_Area_Match']
        df_with_validation.loc[idx, 'GHG_Match'] = validation_df.loc[idx, 'GHG_Match']
        df_with_validation.loc[idx, 'Conditionality_Match'] = validation_df.loc[idx, 'Conditionality_Match']
        df_with_validation.loc[idx, 'Overall_Valid'] = validation_df.loc[idx, 'Overall_Valid']
    
    # Apply visual indicators
    df_with_validation['AI_Is_Target'] = df_with_validation['AI_Is_Target'].map({
        True: '✓ Target',
        False: '❌ Not a target',
        None: 'N/A'
    })
    df_with_validation['Target_Area_Match'] = df_with_validation['Target_Area_Match'].map({
        True: '✓',
        False: '❌ MISMATCH',
        None: 'N/A'
    })
    df_with_validation['GHG_Match'] = df_with_validation['GHG_Match'].map({
        True: '✓',
        False: '❌ MISMATCH',
        None: 'N/A'
    })
    df_with_validation['Conditionality_Match'] = df_with_validation['Conditionality_Match'].map({
        True: '✓',
        False: '❌ MISMATCH',
        None: 'N/A'
    })
    df_with_validation['Overall_Valid'] = df_with_validation['Overall_Valid'].map({
        True: '✓ Valid',
        False: '❌ Invalid',
        None: 'N/A'
    })
    
    # README data
    readme_data = {
        'Sheet Name': [
            'README',
            'Data',
            'Data_with_Flags',
            'Parameter_Validation',
            'False_Positives',
            'Target_Area_Mismatches',
            'GHG_Mismatches',
            'Conditionality_Mismatches'
        ],
        'Description': [
            'This documentation sheet explaining all content - NDC 3.X AI Validation Analysis',
            'NDC 3.X targets dataset (clean, without flags).',
            'NDC 3.X targets with AI validation results.',
            'Summary of AI parameter validation showing match/mismatch statistics.',
            'Entries with Parameter but AI classified as NOT a target (potential false positives).',
            'Rows where AI-predicted Target area differs from database value.',
            'Rows where AI-predicted GHG target classification differs from database value.',
            'Rows where AI-predicted Conditionality differs from database value.'
        ]
    }
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. README
        readme_df = pd.DataFrame(readme_data)
        readme_df.to_excel(writer, sheet_name='README', index=False)
        
        # 2. DATA (clean)
        df_ndc3.to_excel(writer, sheet_name='Data', index=False)
        
        # 3. DATA WITH FLAGS
        df_with_validation.to_excel(writer, sheet_name='Data_with_Flags', index=False)
        
        # 4. PARAMETER VALIDATION SUMMARY
        validation_summary = pd.DataFrame({
            'Metric': [
                'Total NDC 3.X entries validated',
                'AI classified as target',
                'AI classified as NOT target',
                'False positives (has Parameter but not a target)',
                'Target Area matches',
                'Target Area mismatches',
                'GHG target matches',
                'GHG target mismatches',
                'Conditionality matches',
                'Conditionality mismatches',
                'Overall valid entries',
                'Overall invalid entries'
            ],
            'Count': [
                len(validation_df),
                (validation_df['AI_Is_Target'] == True).sum(),
                (validation_df['AI_Is_Target'] == False).sum(),
                validation_df['Is_False_Positive'].sum(),
                (validation_df['Target_Area_Match'] == True).sum(),
                (validation_df['Target_Area_Match'] == False).sum(),
                (validation_df['GHG_Match'] == True).sum(),
                (validation_df['GHG_Match'] == False).sum(),
                (validation_df['Conditionality_Match'] == True).sum(),
                (validation_df['Conditionality_Match'] == False).sum(),
                (validation_df['Overall_Valid'] == True).sum(),
                (validation_df['Overall_Valid'] == False).sum()
            ]
        })
        validation_summary.to_excel(writer, sheet_name='Parameter_Validation', index=False)
        
        # 5. FALSE POSITIVES
        false_positive_indices = validation_df[validation_df['Is_False_Positive'] == True].index
        if len(false_positive_indices) > 0:
            false_positives = df_with_validation.loc[false_positive_indices]
            false_positives.to_excel(writer, sheet_name='False_Positives', index=False)
        
        # 6. TARGET AREA MISMATCHES
        target_area_mismatch_indices = validation_df[validation_df['Target_Area_Match'] == False].index
        if len(target_area_mismatch_indices) > 0:
            target_area_mismatches = df_with_validation.loc[target_area_mismatch_indices]
            priority_cols = ['Target area', 'AI_Target_Area', 'Target_Area_Match', 'Content']
            other_cols = [c for c in target_area_mismatches.columns if c not in priority_cols]
            final_cols = [c for c in priority_cols if c in target_area_mismatches.columns] + other_cols
            target_area_mismatches[final_cols].to_excel(writer, sheet_name='Target_Area_Mismatches', index=False)
        
        # 7. GHG MISMATCHES
        ghg_mismatch_indices = validation_df[validation_df['GHG_Match'] == False].index
        if len(ghg_mismatch_indices) > 0:
            ghg_mismatches = df_with_validation.loc[ghg_mismatch_indices]
            priority_cols = ['GHG target?', 'AI_GHG_Target', 'GHG_Match', 'Content']
            other_cols = [c for c in ghg_mismatches.columns if c not in priority_cols]
            final_cols = [c for c in priority_cols if c in ghg_mismatches.columns] + other_cols
            ghg_mismatches[final_cols].to_excel(writer, sheet_name='GHG_Mismatches', index=False)
        
        # 8. CONDITIONALITY MISMATCHES
        cond_mismatch_indices = validation_df[validation_df['Conditionality_Match'] == False].index
        if len(cond_mismatch_indices) > 0:
            cond_mismatches = df_with_validation.loc[cond_mismatch_indices]
            priority_cols = ['Conditionality', 'AI_Conditionality', 'Conditionality_Match', 'Content']
            other_cols = [c for c in cond_mismatches.columns if c not in priority_cols]
            final_cols = [c for c in priority_cols if c in cond_mismatches.columns] + other_cols
            cond_mismatches[final_cols].to_excel(writer, sheet_name='Conditionality_Mismatches', index=False)
        
        # Apply formatting
        workbook = writer.book
        
        # Add hyperlinks to README
        readme_ws = workbook['README']
        for row_idx in range(2, len(readme_data['Sheet Name']) + 2):
            sheet_name = readme_ws[f'A{row_idx}'].value
            if sheet_name and sheet_name in workbook.sheetnames:
                cell = readme_ws[f'A{row_idx}']
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.font = Font(name='Calibri', size=11, color='0563C1', underline='single')
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Format all sheets
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                fill = alt_fill if row_idx % 2 == 0 else PatternFill()
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.fill = fill
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
            
            ws.freeze_panes = 'A2'
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions
        
        # Special formatting
        readme_ws.column_dimensions['A'].width = 30
        readme_ws.column_dimensions['B'].width = 100
        readme_header_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        for cell in readme_ws[1]:
            cell.fill = readme_header_fill
        
        if 'Parameter_Validation' in workbook.sheetnames:
            val_ws = workbook['Parameter_Validation']
            validation_fill = PatternFill(start_color='9B59B6', end_color='9B59B6', fill_type='solid')
            for cell in val_ws[1]:
                cell.fill = validation_fill
        
        warning_sheets = ['False_Positives', 'Target_Area_Mismatches', 'GHG_Mismatches', 'Conditionality_Mismatches']
        warning_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
        for sheet_name in warning_sheets:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.fill = warning_fill
    
    print(f"\n✓ AI validation report exported to: {output_path}")

if __name__ == "__main__":
    input_file = "dq/excel/NDC-Database-Analysis_current_NEW.xlsx"
    output_file = "dq/ai_validation_ndc-3-X_targets.xlsx"
    
    print("\n" + "="*80)
    print("AI PARAMETER VALIDATION - NDC 3.X TARGETS")
    print("="*80)
    
    # Load data
    print(f"\nLoading data from: {input_file}")
    df = pd.read_excel(input_file, sheet_name='Targets', header=7)
    
    columns_of_interest = [
        'Document ID', 'Country Code', 'Version number', 'Target area',
        'Target scope', 'GHG target?', 'Target type', 'Conditionality',
        'Target Year', 'Content', 'Page Number'
    ]
    available_columns = [col for col in columns_of_interest if col in df.columns]
    df = df[available_columns]
    
    # Filter to NDC 3.X
    ndc3_mask = df['Version number'].str.contains('NDC 3', case=False, na=False)
    df_ndc3 = df[ndc3_mask].copy()
    
    print(f"Loaded {len(df_ndc3)} NDC 3.X target entries")
    
    # Run AI validation
    validation_results = validate_parameter_classification(df, LLM)
    
    if not validation_results.empty:
        # Export results
        export_ai_validation_report(df_ndc3, validation_results, output_file)
        
        print("\n" + "="*80)
        print("AI VALIDATION COMPLETE")
        print("="*80)
    else:
        print("\n⚠️  No validation results to export")