import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

def analyze_excel(file_path):
    """
    Perform basic analysis on an Excel file
    
    Args:
        file_path: Path to the Excel file
    """
    # Read Excel file from 'Adaptation' sheet, starting at row 8 (header at row 7)
    df = pd.read_excel(file_path, sheet_name='Adaptation', header=7)
    
    # Select only the columns of interest for Adaptation analysis
    columns_of_interest = [
        'Document ID',
        'Country Code',
        'Version number',
        'Category',
        'Measure',
        'Quote',
        'Page Number',
        'Activity type'
    ]
    
    # Filter to only include these columns (handles case variations)
    available_columns = [col for col in columns_of_interest if col in df.columns]
    df = df[available_columns]
    
    print(f"Analysis of: {file_path}")
    print("=" * 50)
    
    # Basic information
    print("\n1. Dataset Overview:")
    print(f"   - Rows: {len(df)}")
    print(f"   - Columns: {len(df.columns)}")
    print(f"   - Column names: {list(df.columns)}")
    
    # Data types
    print("\n2. Data Types:")
    print(df.dtypes)
    
    # Missing values
    print("\n3. Missing Values:")
    missing = df.isnull().sum()
    missing_pct = (missing / len(df)) * 100
    missing_df = pd.DataFrame({
        'Missing Count': missing,
        'Percentage': missing_pct
    })
    print(missing_df[missing_df['Missing Count'] > 0])
    
    # Descriptive statistics
    print("\n4. Descriptive Statistics:")
    print(df.describe(include='all'))
    
    # Duplicate rows
    print("\n5. Duplicate Rows:")
    duplicates = df.duplicated().sum()
    print(f"   - Number of duplicates: {duplicates}")
    
    # Value counts for categorical columns
    print("\n6. Value Counts for Key Columns:")
    categorical_cols = ['Country Code', 'Version number', 'Category', 'Measure', 'Activity type']
    for col in categorical_cols:
        if col in df.columns:
            print(f"\n   {col}:")
            value_counts = df[col].value_counts()
            if len(value_counts) > 10:
                print(value_counts.head(10))
                print(f"   ... and {len(value_counts) - 10} more values")
            else:
                print(value_counts)
    
    return df

def export_analysis_report(df, output_path):
    """
    Export analysis results to Excel with multiple sheets
    
    Args:
        df: DataFrame to analyze
        output_path: Path for output Excel file
    """
    # Create a copy with missing value markers
    df_marked = df.copy()
    
    # Add flag columns for missing values
    for col in df.columns:
        flag_col = f'{col}_MISSING'
        df_marked[flag_col] = df[col].isnull().map({True: '❌ MISSING', False: ''})
    
    # README data with sheet names without spaces
    readme_data = {
        'Sheet Name': [
            'README',
            'Data',
            'Data_with_Flags',
            'Statistics',
            'Missing_Values',
            'Duplicate_Rows',
            'Incomplete_Rows',
            'Country_Code',
            'Version_number',
            'Category',
            'Measure',
            'Activity_type',
            'Category-Measure',
            'Activity_by_Category',
            'Measure_by_Country',
            'Category_by_Country'
        ],
        'Description': [
            'This documentation sheet explaining all content',
            'Clean dataset without quality flags (original data only).',
            'Complete dataset with quality flags. Missing values marked with ❌ MISSING.',
            'Descriptive statistics for all columns: count, unique values, top values, frequencies, mean, std, min, max, percentiles.',
            'Summary of missing values per column with count and percentage. Helps identify data completeness issues.',
            'Complete duplicate rows grouped by Duplicate_Group ID. Group numbers start from 1. Review for potential data collection errors.',
            'All rows with at least one missing value, with ❌ flags showing which fields are incomplete.',
            'Value counts for Country Code field showing entries per country.',
            'Value counts for Version number field showing NDC version distribution.',
            'Value counts for Category field showing distribution of adaptation categories.',
            'Value counts for Measure field showing distribution of adaptation measures.',
            'Value counts for Activity type field showing distribution of activity types.',
            'Cross-tabulation showing how many entries exist for each Category-Measure combination.',
            'Cross-tabulation showing Activity type distribution per Category.',
            'Cross-tabulation showing Measure distribution per Country.',
            'Cross-tabulation showing Category distribution per Country.'
        ]
    }
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. README
        readme_df = pd.DataFrame(readme_data)
        readme_df.to_excel(writer, sheet_name='README', index=False)
        
        # 2. DATA
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # 3. DATA WITH FLAGS
        df_marked.to_excel(writer, sheet_name='Data_with_Flags', index=False)
        
        # 4. STATISTICS
        df.describe(include='all').to_excel(writer, sheet_name='Statistics')
        
        # 5. MISSING VALUES
        missing = df.isnull().sum()
        missing_pct = (missing / len(df)) * 100
        missing_df = pd.DataFrame({
            'Column': df.columns,
            'Missing Count': missing.values,
            'Percentage': missing_pct
        })
        missing_df.to_excel(writer, sheet_name='Missing_Values', index=False)
        
        # 6. DUPLICATE ROWS
        duplicates = df[df.duplicated(keep=False)].copy()
        if not duplicates.empty:
            duplicates['Duplicate_Group'] = duplicates.groupby(list(df.columns), dropna=False).ngroup() + 1
            duplicates = duplicates.sort_values(by=['Duplicate_Group'] + list(df.columns))
            cols = ['Duplicate_Group'] + [col for col in df.columns]
            duplicates = duplicates[cols]
            duplicates.to_excel(writer, sheet_name='Duplicate_Rows', index=False)
        
        # 7. INCOMPLETE ROWS
        rows_with_missing = df[df.isnull().any(axis=1)].copy()
        if not rows_with_missing.empty:
            for col in rows_with_missing.columns:
                flag_col = f'{col}_MISSING'
                rows_with_missing[flag_col] = rows_with_missing[col].isnull().map({True: '❌', False: ''})
            rows_with_missing.to_excel(writer, sheet_name='Incomplete_Rows', index=False)
        
        # 8+. VALUE COUNTS
        categorical_cols_mapping = {
            'Country Code': 'Country_Code',
            'Version number': 'Version_number',
            'Category': 'Category',
            'Measure': 'Measure',
            'Activity type': 'Activity_type'
        }
        
        for col, sheet_name in categorical_cols_mapping.items():
            if col in df.columns:
                value_counts = df[col].value_counts(dropna=False).reset_index()
                value_counts.columns = [col, 'Count']
                value_counts.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # CROSS-TABULATIONS
        # Category-Measure relationship
        if 'Category' in df.columns and 'Measure' in df.columns:
            cat_measure = df.groupby(['Category', 'Measure']).size().reset_index(name='Count')
            cat_measure = cat_measure.sort_values(['Category', 'Count'], ascending=[True, False])
            cat_measure.to_excel(writer, sheet_name='Category-Measure', index=False)
        
        # Activity type by Category
        if 'Category' in df.columns and 'Activity type' in df.columns:
            activity_cat = df.groupby(['Category', 'Activity type']).size().reset_index(name='Count')
            activity_cat = activity_cat.sort_values(['Category', 'Count'], ascending=[True, False])
            activity_cat.to_excel(writer, sheet_name='Activity_by_Category', index=False)
        
        # Measure by Country
        if 'Country Code' in df.columns and 'Measure' in df.columns:
            measure_country = df.groupby(['Country Code', 'Measure']).size().reset_index(name='Count')
            measure_country = measure_country.sort_values(['Country Code', 'Count'], ascending=[True, False])
            measure_country.to_excel(writer, sheet_name='Measure_by_Country', index=False)
        
        # Category by Country
        if 'Country Code' in df.columns and 'Category' in df.columns:
            category_country = df.groupby(['Country Code', 'Category']).size().reset_index(name='Count')
            category_country = category_country.sort_values(['Country Code', 'Count'], ascending=[True, False])
            category_country.to_excel(writer, sheet_name='Category_by_Country', index=False)
        
        # Get the workbook to apply formatting and hyperlinks
        workbook = writer.book
        
        # Add hyperlinks to README sheet
        readme_ws = workbook['README']
        for row_idx in range(2, len(readme_data['Sheet Name']) + 2):
            sheet_name = readme_ws[f'A{row_idx}'].value
            if sheet_name and sheet_name in workbook.sheetnames:
                cell = readme_ws[f'A{row_idx}']
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.font = Font(name='Calibri', size=11, color='0563C1', underline='single')
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Format data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                if row_idx % 2 == 0:
                    fill = alt_fill
                else:
                    fill = PatternFill()
                
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.fill = fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze top row
            ws.freeze_panes = 'A2'
            
            # Apply filters to header row
            ws.auto_filter.ref = ws.dimensions
        
        # Special formatting for README sheet
        readme_ws = workbook['README']
        readme_ws.column_dimensions['A'].width = 25
        readme_ws.column_dimensions['B'].width = 100
        
        readme_header_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        for cell in readme_ws[1]:
            cell.fill = readme_header_fill
        
        # Special formatting for error/warning sheets
        warning_sheets = ['Duplicate_Rows', 'Incomplete_Rows']
        warning_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
        
        for sheet_name in warning_sheets:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.fill = warning_fill
        
        # Special formatting for Missing Values sheet
        if 'Missing_Values' in workbook.sheetnames:
            mv_ws = workbook['Missing_Values']
            missing_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
            for cell in mv_ws[1]:
                cell.fill = missing_fill
    
    print(f"\nReport exported to: {output_path}")

if __name__ == "__main__":
    # Specify your Excel file path
    input_file = "dq/excel/NDC-Database-Analysis_current_NEW.xlsx"
    output_file = "dq/analysis_report_adaptation.xlsx"
    
    # Run analysis
    data = analyze_excel(input_file)
    
    # Export report
    export_analysis_report(data, output_file)